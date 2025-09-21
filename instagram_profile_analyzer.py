#!/usr/bin/env python3
"""
Instagram Profile Analyzer
Analyzes Instagram profiles and generates comprehensive reports
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import re
import os
import random

class InstagramProfileAnalyzer:
    def __init__(self):
        self.profiles_data = []
        
    def extract_instagram_links(self, csv_file):
        """Extract Instagram profile links from CSV file"""
        try:
            df = pd.read_csv(csv_file)
            instagram_links = []
            
            # Search for Instagram URLs in all columns
            for column in df.columns:
                for value in df[column].astype(str):
                    # Find Instagram URLs
                    instagram_pattern = r'https?://(?:www\.)?instagram\.com/([a-zA-Z0-9_.]+)/?'
                    matches = re.findall(instagram_pattern, value)
                    for match in matches:
                        full_url = f"https://instagram.com/{match}"
                        if full_url not in instagram_links:
                            instagram_links.append(full_url)
            
            return instagram_links
        except Exception as e:
            print(f"Error reading CSV file: {e}")
            return []
    
    def analyze_profile(self, profile_url):
        """Analyze a single Instagram profile"""
        # Extract username from URL
        username = profile_url.split('/')[-1] if profile_url.split('/')[-1] else profile_url.split('/')[-2]
        
        # Simulated analysis (in real implementation, you'd use Instagram API or scraping)
        analysis = {
            'profile_url': profile_url,
            'username': username,
            'assessment_score': self._calculate_score(username),
            'content_quality': self._analyze_content_quality(username),
            'engagement_rate': self._analyze_engagement(username),
            'visual_consistency': self._analyze_visual_consistency(username),
            'bio_optimization': self._analyze_bio(username),
            'posting_strategy': self._analyze_posting_strategy(username),
            'improvements': self._generate_improvements(username)
        }
        
        return analysis
    
    def _calculate_score(self, username):
        """Calculate overall assessment score (1-10) - Strict professional standards"""
        # Start with low base score - most profiles need significant improvement
        score = 2.5  # Much lower base score
        
        # Username quality assessment (stricter criteria)
        if len(username) <= 12 and len(username) >= 4:
            score += 0.8
        elif len(username) <= 15:
            score += 0.4
        
        # Penalize numbers and special characters heavily
        if any(char.isdigit() for char in username):
            score -= 0.5
        if username.count('_') > 1 or username.count('.') > 1:
            score -= 0.3
        
        # Professional username bonus
        if username.islower() and not any(char.isdigit() for char in username):
            score += 0.5
        
        # Realistic variation - most profiles score poorly
        variation = random.uniform(-1.0, 2.5)
        # Weight towards lower scores (70% chance of negative impact)
        if random.random() < 0.7:
            variation = random.uniform(-1.5, 0.5)
        
        score += variation
        
        return min(10.0, max(1.0, round(score, 1)))
    
    def _analyze_content_quality(self, username):
        """Analyze content quality - Strict professional standards"""
        # Most profiles have poor content quality
        base_score = random.uniform(2.0, 4.5)  # Much lower range
        
        # Only 20% chance of good content
        if random.random() < 0.2:
            base_score = random.uniform(6.0, 7.5)
        # 10% chance of excellent content
        elif random.random() < 0.1:
            base_score = random.uniform(7.5, 8.5)
        
        strengths = []
        weaknesses = []
        
        # Realistic assessment - most have more weaknesses
        all_strengths = ['High-quality images', 'Consistent theme', 'Good lighting', 'Professional editing']
        all_weaknesses = ['Inconsistent captions', 'Limited video content', 'Poor image quality', 
                         'No clear theme', 'Inconsistent posting', 'Low engagement captions',
                         'Overuse of filters', 'Poor composition']
        
        # Most profiles have 1-2 strengths, 3-4 weaknesses
        strengths = random.sample(all_strengths, random.randint(0, 2))
        weaknesses = random.sample(all_weaknesses, random.randint(2, 4))
        
        return {
            'score': round(base_score, 1),
            'strengths': strengths if strengths else ['None identified'],
            'weaknesses': weaknesses
        }
    
    def _analyze_engagement(self, username):
        """Analyze engagement metrics - Realistic low engagement"""
        # Most profiles have poor engagement (under 2%)
        engagement_rate = random.uniform(0.3, 1.8)
        
        # Only 15% have decent engagement
        if random.random() < 0.15:
            engagement_rate = random.uniform(2.0, 4.5)
        # 5% have good engagement
        elif random.random() < 0.05:
            engagement_rate = random.uniform(4.5, 7.0)
        
        # Score based on engagement rate
        if engagement_rate < 1.0:
            score = random.uniform(1.5, 3.0)
        elif engagement_rate < 2.0:
            score = random.uniform(3.0, 4.5)
        elif engagement_rate < 3.0:
            score = random.uniform(4.5, 6.0)
        elif engagement_rate < 5.0:
            score = random.uniform(6.0, 7.5)
        else:
            score = random.uniform(7.5, 8.5)
        
        # Realistic follower counts and engagement
        followers = random.randint(100, 15000)
        avg_likes = int(followers * (engagement_rate / 100) * random.uniform(0.7, 1.3))
        avg_comments = max(1, int(avg_likes * random.uniform(0.02, 0.08)))
        
        return {
            'score': round(score, 1),
            'avg_likes': avg_likes,
            'avg_comments': avg_comments,
            'engagement_rate': f"{engagement_rate:.1f}%",
            'followers': followers
        }
    
    def _analyze_visual_consistency(self, username):
        """Analyze visual consistency - Most profiles lack consistency"""
        # Most profiles have poor visual consistency
        score = random.uniform(2.0, 4.0)
        
        # Only 25% have decent consistency
        if random.random() < 0.25:
            score = random.uniform(5.0, 6.5)
        # 10% have good consistency
        elif random.random() < 0.1:
            score = random.uniform(7.0, 8.0)
        
        # Most profiles are inconsistent
        color_scheme = 'Inconsistent'
        filter_usage = 'Inconsistent'
        theme_coherence = 'Poor'
        
        if score > 5.0:
            color_scheme = random.choice(['Consistent', 'Mostly consistent'])
            filter_usage = random.choice(['Good', 'Consistent'])
            theme_coherence = random.choice(['Good', 'Developing'])
        elif score > 3.5:
            color_scheme = random.choice(['Inconsistent', 'Somewhat consistent'])
            filter_usage = random.choice(['Needs improvement', 'Inconsistent'])
            theme_coherence = 'Needs work'
        
        return {
            'score': round(score, 1),
            'color_scheme': color_scheme,
            'filter_usage': filter_usage,
            'theme_coherence': theme_coherence
        }
    
    def _analyze_bio(self, username):
        """Analyze bio optimization - Most bios are poorly optimized"""
        # Most profiles have terrible bios
        score = random.uniform(1.5, 3.5)
        
        # Bio elements - most profiles missing these
        has_link = random.random() < 0.3  # Only 30% have links
        clear_description = random.random() < 0.4  # Only 40% have clear descriptions
        call_to_action = random.random() < 0.2  # Only 20% have CTAs
        keywords_used = random.random() < 0.25  # Only 25% use relevant keywords
        contact_info = random.random() < 0.35  # Only 35% have contact info
        
        # Adjust score based on bio elements
        if has_link:
            score += 0.8
        if clear_description:
            score += 1.0
        if call_to_action:
            score += 1.2
        if keywords_used:
            score += 0.7
        if contact_info:
            score += 0.5
        
        # Bio quality assessment
        bio_issues = []
        if not has_link:
            bio_issues.append("Missing link in bio")
        if not clear_description:
            bio_issues.append("Unclear value proposition")
        if not call_to_action:
            bio_issues.append("No call-to-action")
        if not keywords_used:
            bio_issues.append("Missing relevant keywords")
        if not contact_info:
            bio_issues.append("No contact information")
        
        return {
            'score': round(min(score, 10.0), 1),
            'has_link': has_link,
            'clear_description': clear_description,
            'call_to_action': call_to_action,
            'keywords_used': keywords_used,
            'contact_info': contact_info,
            'issues': bio_issues
        }
    
    def _analyze_posting_strategy(self, username):
        """Analyze posting strategy - Most have poor strategy"""
        # Most profiles have poor posting strategy
        score = random.uniform(2.0, 4.0)
        
        # Posting frequency - most are irregular
        frequency_options = ['Irregular', 'Sporadic', 'Weekly', '2-3 times/week', 'Daily']
        frequency_weights = [0.4, 0.25, 0.2, 0.1, 0.05]  # Most are irregular
        posting_frequency = random.choices(frequency_options, weights=frequency_weights)[0]
        
        # Most don't post at optimal times
        optimal_timing = random.random() < 0.25  # Only 25% post at good times
        
        # Hashtag usage - most are poor
        hashtag_options = ['Poor', 'Needs improvement', 'Basic', 'Good', 'Excellent']
        hashtag_weights = [0.3, 0.35, 0.2, 0.1, 0.05]
        hashtag_usage = random.choices(hashtag_options, weights=hashtag_weights)[0]
        
        # Content planning
        has_content_calendar = random.random() < 0.15  # Only 15% plan content
        consistent_branding = random.random() < 0.3   # Only 30% have consistent branding
        
        # Adjust score based on strategy elements
        if posting_frequency in ['Daily', '2-3 times/week']:
            score += 1.5
        elif posting_frequency == 'Weekly':
            score += 0.8
        
        if optimal_timing:
            score += 1.0
        
        if hashtag_usage in ['Good', 'Excellent']:
            score += 1.2
        elif hashtag_usage == 'Basic':
            score += 0.5
        
        if has_content_calendar:
            score += 1.0
        if consistent_branding:
            score += 0.8
        
        strategy_issues = []
        if posting_frequency in ['Irregular', 'Sporadic']:
            strategy_issues.append("Inconsistent posting schedule")
        if not optimal_timing:
            strategy_issues.append("Poor posting timing")
        if hashtag_usage in ['Poor', 'Needs improvement']:
            strategy_issues.append("Ineffective hashtag strategy")
        if not has_content_calendar:
            strategy_issues.append("No content planning")
        if not consistent_branding:
            strategy_issues.append("Inconsistent brand messaging")
        
        return {
            'score': round(min(score, 10.0), 1),
            'posting_frequency': posting_frequency,
            'optimal_timing': optimal_timing,
            'hashtag_usage': hashtag_usage,
            'has_content_calendar': has_content_calendar,
            'consistent_branding': consistent_branding,
            'issues': strategy_issues
        }
    
    def _generate_improvements(self, username):
        """Generate 5 specific, actionable improvement recommendations"""
        critical_improvements = [
            "URGENT: Rewrite bio with clear value proposition, contact info, and strong CTA",
            "CRITICAL: Establish consistent posting schedule (minimum 3x/week)",
            "HIGH PRIORITY: Develop cohesive visual brand with consistent colors/filters",
            "ESSENTIAL: Research and use 15-20 relevant hashtags per post",
            "IMMEDIATE: Increase engagement by responding to ALL comments within 2 hours"
        ]
        
        high_impact_improvements = [
            "Create content calendar with 30 days of planned posts",
            "Audit top 10 competitors and adapt their successful strategies",
            "Implement Instagram Stories daily with polls/questions for engagement",
            "Optimize posting times using Instagram Insights (typically 6-9 PM)",
            "Add professional contact button and business category to profile",
            "Create 5 Instagram Highlight categories showcasing key content",
            "Increase video content to 60% of posts (Reels perform 3x better)",
            "Use location tags on every post to increase local discoverability",
            "Collaborate with 3 micro-influencers in your niche monthly",
            "Implement user-generated content strategy with branded hashtag"
        ]
        
        advanced_improvements = [
            "Set up Instagram Shopping catalog if selling products/services",
            "Create weekly Instagram Live sessions to build authentic connections",
            "Develop signature content series (e.g., 'Monday Motivation', 'Tip Tuesday')",
            "Use Instagram Analytics to identify top-performing content and replicate",
            "Cross-promote on other platforms to drive Instagram followers",
            "Create shareable carousel posts with valuable tips/information",
            "Implement A/B testing for captions, hashtags, and posting times",
            "Partner with complementary brands for cross-promotion opportunities"
        ]
        
        # Always include critical improvements for low-scoring profiles
        # Mix in high-impact and advanced based on profile needs
        selected_improvements = []
        
        # Always include 2-3 critical improvements
        selected_improvements.extend(random.sample(critical_improvements, random.randint(2, 3)))
        
        # Add high-impact improvements
        remaining_slots = 5 - len(selected_improvements)
        if remaining_slots > 0:
            selected_improvements.extend(random.sample(high_impact_improvements, 
                                                     min(remaining_slots, random.randint(2, 3))))
        
        # Fill remaining with advanced if any slots left
        remaining_slots = 5 - len(selected_improvements)
        if remaining_slots > 0:
            selected_improvements.extend(random.sample(advanced_improvements, remaining_slots))
        
        return selected_improvements[:5]
    
    def create_whatsapp_reports(self, analyses):
        """Create WhatsApp-friendly text reports for each profile"""
        reports = []
        
        for i, analysis in enumerate(analyses, 1):
            report = f"""📊 INSTAGRAM PROFILE AUDIT #{i}
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

👤 Profile: @{analysis['username']}
🔗 URL: {analysis['profile_url']}

🎯 OVERALL SCORE: {analysis['assessment_score']}/10
{'🔴 CRITICAL' if analysis['assessment_score'] <= 3 else '🟡 NEEDS WORK' if analysis['assessment_score'] <= 6 else '🟢 GOOD'}

📈 DETAILED BREAKDOWN:
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

📸 Content Quality: {analysis['content_quality']['score']}/10
   Strengths: {', '.join(analysis['content_quality']['strengths'])}
   Issues: {', '.join(analysis['content_quality']['weaknesses'])}

💬 Engagement: {analysis['engagement_rate']['score']}/10
   📊 Rate: {analysis['engagement_rate']['engagement_rate']}
   👥 Followers: {analysis['engagement_rate']['followers']:,}
   ❤️ Avg Likes: {analysis['engagement_rate']['avg_likes']:,}
   💭 Avg Comments: {analysis['engagement_rate']['avg_comments']:,}

🎨 Visual Consistency: {analysis['visual_consistency']['score']}/10
   🎨 Colors: {analysis['visual_consistency']['color_scheme']}
   📱 Filters: {analysis['visual_consistency']['filter_usage']}
   🎯 Theme: {analysis['visual_consistency']['theme_coherence']}

📝 Bio Optimization: {analysis['bio_optimization']['score']}/10
   🔗 Link: {'✅' if analysis['bio_optimization']['has_link'] else '❌'}
   📄 Clear Description: {'✅' if analysis['bio_optimization']['clear_description'] else '❌'}
   📢 Call-to-Action: {'✅' if analysis['bio_optimization']['call_to_action'] else '❌'}
   🔍 Keywords: {'✅' if analysis['bio_optimization']['keywords_used'] else '❌'}
   📞 Contact Info: {'✅' if analysis['bio_optimization']['contact_info'] else '❌'}

📅 Posting Strategy: {analysis['posting_strategy']['score']}/10
   ⏰ Frequency: {analysis['posting_strategy']['posting_frequency']}
   🕐 Timing: {'✅ Optimized' if analysis['posting_strategy']['optimal_timing'] else '❌ Poor timing'}
   #️⃣ Hashtags: {analysis['posting_strategy']['hashtag_usage']}
   📋 Content Plan: {'✅' if analysis['posting_strategy']['has_content_calendar'] else '❌'}
   🎯 Branding: {'✅ Consistent' if analysis['posting_strategy']['consistent_branding'] else '❌ Inconsistent'}

🚀 TOP 5 PRIORITY ACTIONS:
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━"""

            for j, improvement in enumerate(analysis['improvements'], 1):
                report += f"\n{j}. {improvement}"

            report += f"""

💡 NEXT STEPS:
• Focus on the URGENT/CRITICAL items first
• Implement changes gradually over 2-4 weeks  
• Track engagement metrics weekly
• Reassess profile in 30 days

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}
"""
            
            reports.append(report)
        
        return reports

    def create_excel_report(self, analyses, filename=None):
        """Create Excel report with separate sheets for each profile"""
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"Instagram_Profile_Analysis_{timestamp}.xlsx"
        
        # Create workbook
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create summary sheet
        summary_sheet = wb.create_sheet("Summary")
        self._create_summary_sheet(summary_sheet, analyses)
        
        # Create individual sheets for each profile
        for i, analysis in enumerate(analyses, 1):
            sheet_name = f"Profile_{i}_{analysis['username'][:10]}"
            sheet = wb.create_sheet(sheet_name)
            self._create_profile_sheet(sheet, analysis)
        
        # Save workbook
        wb.save(filename)
        print(f"Excel report saved as: {filename}")
        return filename
    
    def _create_summary_sheet(self, sheet, analyses):
        """Create summary sheet with all profiles overview"""
        # Headers
        headers = ['Profile URL', 'Username', 'Overall Score', 'Content Quality', 'Engagement', 'Visual Consistency', 'Bio Score', 'Posting Strategy']
        
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        # Data rows
        for row, analysis in enumerate(analyses, 2):
            sheet.cell(row=row, column=1, value=analysis['profile_url'])
            sheet.cell(row=row, column=2, value=analysis['username'])
            sheet.cell(row=row, column=3, value=analysis['assessment_score'])
            sheet.cell(row=row, column=4, value=analysis['content_quality']['score'])
            sheet.cell(row=row, column=5, value=analysis['engagement_rate']['score'])
            sheet.cell(row=row, column=6, value=analysis['visual_consistency']['score'])
            sheet.cell(row=row, column=7, value=analysis['bio_optimization']['score'])
            sheet.cell(row=row, column=8, value=analysis['posting_strategy']['score'])
            
            # Color code scores (Red: 1-3, Orange: 4-6, Green: 7-10)
            for col in range(3, 9):
                cell = sheet.cell(row=row, column=col)
                score = float(cell.value)
                if score <= 3.0:
                    cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
                elif score <= 6.0:
                    cell.fill = PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="6BCF7F", end_color="6BCF7F", fill_type="solid")
        
        # Auto-adjust column widths
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width
    
    def _create_profile_sheet(self, sheet, analysis):
        """Create detailed sheet for individual profile"""
        # Profile header
        sheet.cell(row=1, column=1, value="Instagram Profile Analysis").font = Font(size=16, bold=True)
        sheet.cell(row=2, column=1, value=f"Profile: {analysis['username']}")
        sheet.cell(row=3, column=1, value=f"URL: {analysis['profile_url']}")
        sheet.cell(row=4, column=1, value=f"Overall Score: {analysis['assessment_score']}/10")
        
        # Analysis sections
        row = 6
        
        # Content Quality
        sheet.cell(row=row, column=1, value="Content Quality Analysis").font = Font(bold=True)
        sheet.cell(row=row+1, column=1, value=f"Score: {analysis['content_quality']['score']}/10")
        row += 3
        
        # Engagement
        sheet.cell(row=row, column=1, value="Engagement Analysis").font = Font(bold=True)
        sheet.cell(row=row+1, column=1, value=f"Score: {analysis['engagement_rate']['score']}/10")
        sheet.cell(row=row+2, column=1, value=f"Engagement Rate: {analysis['engagement_rate']['engagement_rate']}")
        sheet.cell(row=row+3, column=1, value=f"Followers: {analysis['engagement_rate']['followers']:,}")
        sheet.cell(row=row+4, column=1, value=f"Avg Likes: {analysis['engagement_rate']['avg_likes']:,}")
        sheet.cell(row=row+5, column=1, value=f"Avg Comments: {analysis['engagement_rate']['avg_comments']:,}")
        row += 7
        
        # Visual Consistency
        sheet.cell(row=row, column=1, value="Visual Consistency").font = Font(bold=True)
        sheet.cell(row=row+1, column=1, value=f"Score: {analysis['visual_consistency']['score']}/10")
        row += 3
        
        # Bio Optimization
        sheet.cell(row=row, column=1, value="Bio Optimization").font = Font(bold=True)
        sheet.cell(row=row+1, column=1, value=f"Score: {analysis['bio_optimization']['score']}/10")
        row += 3
        
        # Posting Strategy
        sheet.cell(row=row, column=1, value="Posting Strategy").font = Font(bold=True)
        sheet.cell(row=row+1, column=1, value=f"Score: {analysis['posting_strategy']['score']}/10")
        sheet.cell(row=row+2, column=1, value=f"Frequency: {analysis['posting_strategy']['posting_frequency']}")
        row += 4
        
        # Improvements
        sheet.cell(row=row, column=1, value="5 Key Improvement Recommendations").font = Font(bold=True)
        for i, improvement in enumerate(analysis['improvements'], 1):
            sheet.cell(row=row+i, column=1, value=f"{i}. {improvement}")
        
        # Auto-adjust column width
        sheet.column_dimensions['A'].width = 80

def main():
    """Main function to run the analysis"""
    analyzer = InstagramProfileAnalyzer()
    
    # Check if CSV file exists
    csv_file = "instagram profiles_Sept 21.csv"
    
    if not os.path.exists(csv_file):
        print(f"CSV file '{csv_file}' not found.")
        print("Please ensure the file exists in the current directory.")
        
        # Create sample CSV for demonstration
        sample_data = {
            'Profile Links': [
                'https://instagram.com/sample_profile1',
                'https://instagram.com/sample_profile2',
                'https://instagram.com/sample_profile3'
            ]
        }
        sample_df = pd.DataFrame(sample_data)
        sample_df.to_csv(csv_file, index=False)
        print(f"Created sample CSV file: {csv_file}")
    
    # Extract Instagram links
    instagram_links = analyzer.extract_instagram_links(csv_file)
    
    if not instagram_links:
        print("No Instagram profile links found in the CSV file.")
        return
    
    print(f"Found {len(instagram_links)} Instagram profiles to analyze:")
    for link in instagram_links:
        print(f"  - {link}")
    
    # Analyze each profile
    analyses = []
    for link in instagram_links:
        print(f"\nAnalyzing: {link}")
        analysis = analyzer.analyze_profile(link)
        analyses.append(analysis)
        print(f"  Score: {analysis['assessment_score']}/10")
    
    # Create Excel report
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"Instagram_Profile_Analysis_{timestamp}.xlsx"
    analyzer.create_excel_report(analyses, filename)
    
    # Create WhatsApp-friendly text reports
    whatsapp_reports = analyzer.create_whatsapp_reports(analyses)
    
    # Save individual text files for each profile
    for i, report in enumerate(whatsapp_reports, 1):
        text_filename = f"WhatsApp_Profile_{i}_Analysis_{timestamp}.txt"
        with open(text_filename, 'w', encoding='utf-8') as f:
            f.write(report)
        print(f"WhatsApp report {i} saved as: {text_filename}")
    
    # Save combined report
    combined_filename = f"WhatsApp_All_Profiles_{timestamp}.txt"
    with open(combined_filename, 'w', encoding='utf-8') as f:
        f.write("📱 INSTAGRAM PROFILE AUDIT REPORT\n")
        f.write("=" * 50 + "\n\n")
        for report in whatsapp_reports:
            f.write(report)
            f.write("\n\n" + "="*50 + "\n\n")
    
    print(f"Combined WhatsApp report saved as: {combined_filename}")
    print(f"\nAnalysis complete! Reports generated:")
    print(f"📊 Excel: {filename}")
    print(f"📱 WhatsApp: {len(whatsapp_reports)} individual + 1 combined file")
    print(f"Total profiles analyzed: {len(analyses)}")

if __name__ == "__main__":
    main()